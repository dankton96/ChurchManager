from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime,platform,os,time,easygui

##########################################################################################################
def isLineNone(fline,fNoneCounter):
    fNoneCounter=0
    for i in fline:
        if i.value==None:
            fNoneCounter+=1
    return fNoneCounter==len(fline)
##########################################################################################################
def isPageEmpty(xlpage0):
    nonect=0
    if(xlpage0.max_row==1):
        return isLineNone(xlpage0[xlpage0.max_row],nonect)
    else:
        noneLines=0
        for i in xlpage0:
            nonect=0
            if(isLineNone(i,nonect)):noneLines+=1
        return noneLines==xlpage0.max_row
##########################################################################################################
def RemoveEmptyRows(xlpage):
    lineIndex=0
    lineToRemove=[]
    for line in xlpage:
        NoneCounter=0
        lineIndex+=1
        if isLineNone(line,NoneCounter):
            lineToRemove.append(lineIndex)
        else:
            if NoneCounter>0:
                return "RegInconsistencyError",lineIndex
    for i in reversed(lineToRemove):
        xlpage.delete_rows(i)
    return "RemotionSuccessful",0
##########################################################################################################
def isLeapYear(year):
    if(year%4==0):
        return (year%100)!=0
    else:
        return (year%400)==0
##########################################################################################################
def isDateValid(date):
    #date in format [d,m,y]
    NormalYear={1:30,2:28,3:30,4:31,5:30,6:31,7:30,8:30,9:31,10:30,11:31,12:30}
    LeapYear={1:30,2:29,3:30,4:31,5:30,6:31,7:30,8:30,9:31,10:30,11:31,12:30}
    if(date[1] not in range(1,13)):
        return False
    if(isLeapYear(date[2])):
        if(date[0]>LeapYear[date[1]]):
            return False
    else:
        if(date[0]>NormalYear[date[1]]):
            return False
    return True
##########################################################################################################
def AddUser(reg,row):
    #verify if the data that the user is trying to add is valid and without inconsistency
    if(len(row)!=3): return "InvalidRegError"
    for i in row:
        if(i==None): return "InvalidRegError"
    if(not(isDateValid(row[2]))):
        return "InvalidDateError"
    if(not(isPageEmpty(reg['Usuarios']))):
        existingReg=[i[1].value for i in reg['Usuarios']]
        if(row[1] in existingReg): return "UserAlreadyCadError"
    date=datetime.date(row[2][2],row[2][1],row[2][0])
    reg['Usuarios'].append([row[0],row[1],date])
    #create the page of the user (which contains the user payments)
    newsheetname="Pg_User"+str(row[1])
    reg.create_sheet(newsheetname)
##########################################################################################################
def DelUser(reg,row):
    #verify if the data that the user is trying to add is valid and without inconsistency
    if(len(row)!=3): return "InvalidRegError"
    for i in row:
        if(i==None): return "InvalidRegError"
    if(not(isDateValid(row[2]))):
        return "InvalidDateError"
    date=datetime.date(row[2][2],row[2][1],row[2][0])
    if([row[0],row[1],date]==reg[row[1]]):
        reg['Usuarios'].delete_rows(row[1])
        #delete the page of the user (which contains the user payments)
        sheetToDelName="Pg_User"+str(row[1])
        reg.remove_sheet(sheetToDelName)
        return "RemotionSuccess"
    else:
        if([row[0],row[1],date]not in reg):
            return "UserNotRegistered"
        else:
            return "InconsistentDB_Error"
##########################################################################################################
def ClearNoneInWorkbook(wkbook2clPath):
    wkbook2cl=load_workbook(wkbook2clPath)
    for page in wkbook2cl:
        print(page)
        if(not(isPageEmpty(page))):
            BDFix,Mline=RemoveEmptyRows(page)
            print(BDFix)
    wkbook2cl.save(wkbook2clPath)
##########################################################################################################
def ClearScreen():
    if(platform.system()=="Windows"): os.system("cls")
    if(platform.system()=="Linux"): os.system("clear")
##########################################################################################################
def CheckHeadersUsers(database):
    return (database["Usuarios"]['A1']=="Nome" and database["Usuarios"]['B1']=="Matricula" and database["Usuarios"]['C1']=="DataMatricula")       
##########################################################################################################
def CheckHeadersPersonalPage(page):
    return (page['A1']=="Valor" and page['B1']=="Vencimento" and page['C1']=="ValorPago")
##########################################################################################################
def InitBD(reg,filename):
    ct=0
    for page in reg:
        if(ct==0):
                if(page.title!="Usuarios"):
                    page.title="Usuarios"
                    page['A1']="Nome"
                    page['B1']="Matricula"
                    page['C1']="DataMatricula"
                    ct+=1
        else:
            if(page.title=="Usuarios"):
                    if(not(CheckHeadersUsers(reg))):
                        page['A1']="Nome"
                        page['B1']="Matricula"
                        page['C1']="DataMatricula"
            else:
                if("Pg_User" in page.title):
                    if(not(CheckHeadersPersonalPage(page))):
                       page['A1']="Valor"
                       page['B1']="Vencimento"
                       page['C1']="ValorPago"
    ClearNoneInWorkbook(filename)
    reg.save(filename)
##########################################################################################################
def GetNewRegCode(bdname):
    useBd=load_workbook(bdname)
    return useBd['Usuarios'].max_row
##########################################################################################################
def FindCad(method,bd,toFind):
    Found=[]
    i=0
    for line in bd['Usuarios']:
        if(i!=0):
            if(method=='1'):
                if(toFind in line[0].value):
                    Found.append([line[0],line[1],line[2]])
            if(method=='2'):
                if(int(line[1].value)==toFind):
                    Found.append([line[0],line[1],line[2]])
            if(method=='3'):
                if(line[2]==toFind):
                    Found.append([line[0],line[1],line[2]])
        i+=1
    return Found
##########################################################################################################
def isStrADate(dateTxt):#dd/mm/aaaa
    cnum=['0','1','2','3','4','5','6','7','8','9']
    fct=0
    for c in date:
        if(fct in [2,5]):
            if(c not in ['/','-']):
                fct+=1
                return False
        else:
            if(c not in cnum):
                fct+=1
                return False
    return True
##########################################################################################################
rel={"January":1,"February":2,"March":3,"April":4,"May":5,"June":6,"July":7,"August":8,"September":9,"October":10,"November":11,"December":12}
menuTxt="""
1)Adicionar novo cadastro
2)Remover um cadastro existente
3)Consultar cadastro
4)Registrar pagamento de dízimo
5)Gerar carteira de fiel
6)Sair
"""
searchOP="""
1)Filtrar por nome
2)Filtrar por matrícula
3)Filtrar por data de matrícula
"""
ValidOP=['1','2','3','4','5','6']
searchMethods=['1','2','3']
def Menu(bdfile,path):
    InitBD(bdfile,path)
    while(True):
        ClearScreen()
        op=0
        InitBD(bdfile,path)
        while(op not in ValidOP):
            ClearScreen()
            op=input(menuTxt)
        if(op=='1'):
            name=input("Nome:")
            y=int(time.strftime("%Y"))
            m=rel[time.strftime("%B")]
            d=int(time.strftime("%d"))
            code=GetNewRegCode(path)
            AddUser(bdfile,[name,code,[d,m,y]])
            ClearNoneInWorkbook(path)
        if(op=='2'):
            findMethod=0
            while(findMethod not in searchMethods):
                ClearScreen()
                findMethod=input(searchOP)
            if(findMethod=='1'):
                target=input("Insira o nome do usuário a ser removido:")
            if(findMethod=='2'):
                target=input("Insira a matrícula do usuário a ser removido:")
            if(findMethod=='3'):
                target=input("Insira a data de matrícula do usuário a ser removido:")
            ToDel=FindCad(findMethod,bdfile,target)
            if(len(ToDel)>1):
                SelectedReg=0
                validOp=[str(c) for c in range(1,(len(ToDel)+1))]
                while(SelectedReg not in validOp):
                    ClearScreen()
                    nf=str(len(ToDel))
                    print("Foram encontrados mais de um usuário com o critério informado. Selecione o que deseja remover:\n")
                    print(" ".ljust((len(nf)+4)),"{:20}{:12}{:10}".format("Nome","Matrícula","Data de matrícula"))
                    print("-"*(50+len(nf)+5))
                    ct=0
                    for i in ToDel:
                        ct+=1
                        dtxt=str(i[2].value)
                        data=datetime.date(int(dtxt[0:4]),int(dtxt[5:7]),int(dtxt[8:10]))
                        print((str(ct)+"- ").rjust(len(nf),'0'),"{:15}{:12}".format(i[0].value,i[1].value),"{:>20}".format(str(data)))
                        #ljust(a1,a2) put the content more to left side of the reserved space, a1 is the space that will be, a2 is what will fill the unused space
                        #rjust(a1,a2) put it more to right side, a1 is the space that will be, a2 is what will fill the unused space
                    SelectedReg=input()
                #nome, matricula, data
                DelUser(bdfile,[ToDel[SelectedReg][0],ToDel[SelectedReg][1],dtxt])
            else:
                dtxt=str(ToDel[2].value)
                data=datetime.date(int(dtxt[0:4]),int(dtxt[5:7]),int(dtxt[8:10]))
                DelUser(bdfile,[ToDel[0],ToDel[1],dtxt])
                
        if(op=='3'):
            findMethod=0
            while(findMethod not in searchMethods):
                ClearScreen()
                findMethod=input(searchOP)
            if(findMethod=='1'):
                target=input("Insira o nome a ser localizado:")
            if(findMethod=='2'):
                target=input("Insira a matrícula a ser localizada:")
            if(findMethod=='3'):
                target=input("Insira a data de matrícula a ser localizada:")
            res=FindCad(findMethod,bdfile,target)
            print("{:20}{:12}{:10}".format("Nome","Matrícula","Data de matrícula"))
            print("-"*50)
            for i in res:
                data=str(i[2].value)
                print("{:20}{:<12}{:10}".format(i[0].value,i[1].value,data[:10]))
            var=input("\n\nPressione enter para continuar")
        if(op=='4'):
            print("Nao implementada ainda")
            time.sleep(5)
        if(op=='5'):
            print("Nao implementada ainda")
            time.sleep(5)
        if(op=='6'):
            bdfile.save(path)
            break
file=easygui.fileopenbox("Selecione o arquivo XLSX a ser manipulado:",default="*.xlsx")
fileToUse=load_workbook(file)
Menu(fileToUse,file)
